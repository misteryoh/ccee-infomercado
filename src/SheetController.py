import openpyxl
import json
from io import StringIO


class SheetData:

    def __new__(cls, *args, **kwargs):
        return super().__new__(cls)
    
    def __init__(self, workbook=None):
        self.workbook = workbook

    def load_workbook(self, filepath, filename):
        self.filepath = filepath
        self.filename = filename
        self.workbook = openpyxl.load_workbook(filepath + "/" + filename, read_only=True)
        return self
    
    def extract_data(self, params, *args):
        """Extract structured data from downloaded file from CCEE
        :param params: {sheet_name, table_name, first_col, last_col, footer, deadrows}
        """

        results = []

        for sheets in params:

            sheet = self.workbook[sheets['sheet_name']]
            title_cell = sheets['table_name']
            first_cell = sheets['first_col']
            last_cell = sheets['last_col']
            footer_cell = sheets['footer']
            deadrows = sheets['deadrows']
            output_name = sheets['output_name']
            output_type = sheets['output_type']
            schema = sheets['schema']

            title_range = None
            start_range = None
            last_range = None
            foot_range = None

            for row in sheet.iter_rows():
                for cell in row:
                    if title_cell in str(cell.value):
                        title_range = cell
                    if first_cell in str(cell.value):
                        start_range = cell
                    if last_cell in str(cell.value):
                        last_range = cell
                    if footer_cell in str(cell.value):
                        foot_range = cell
                        break

            print("title_range  : " + title_range.coordinate)
            print("start_range  : " + start_range.coordinate)
            print("last_range  : " + last_range.coordinate)
            print("foot_range   : " + foot_range.coordinate)


            # Load schema
            with open(f'/home/misteryoh/Coding/Git/ccee-infomercado/data/{schema}', 'r') as schema_file:
                schema_data = json.load(schema_file)
                colunas_originais = schema_data["colunas_originais"]
                colunas_novas = schema_data["colunas_novas"]

            # Criação do dicionário para mapear colunas originais para novas colunas
            colunas_mapping = dict(zip(colunas_originais, colunas_novas))

            csv_buffer = StringIO()
            count_row = 0

            if title_range.value is not None:
                for row in sheet[f'{start_range.coordinate}':f'{last_range.column_letter + str(foot_range.row - deadrows)}']:
                    for cell in row:
                        if count_row == 0:
                            valor_celula = str(cell.value)
                            valor_coluna_nova = colunas_mapping.get(valor_celula, valor_celula)

                            csv_buffer.write(valor_coluna_nova + ';')
                        else:
                            csv_buffer.write(str(cell.value) + ';')
                    csv_buffer.write('\n')
                    count_row += 1
            else:
                print("Algo deu errado")

            csv_buffer.seek(0)

            sheet_data = {}
            sheet_data['file'] = output_name
            sheet_data['data'] = csv_buffer.getvalue().encode()
            sheet_data['type'] = output_type

            results.append(sheet_data)

        return results